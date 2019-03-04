import cognitive_face as CF
from global_variables import personGroupId
import os, urllib
import sqlite3
from openpyxl import Workbook, load_workbook

try:
    from openpyxl.cell import get_column_letter, column_index_from_string
except ImportError:
    from openpyxl.utils import get_column_letter, column_index_from_string
import time

# get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename="reports.xlsx")
sheet = wb.get_sheet_by_name('16CE')


def getDateColumn():
    for i in range(1,sheet.get_highest_row() + 1):
        col = get_column_letter(i)
        if sheet.cell('%s%s' % (col, '1')).value == currentDate:
            return col


Key = 'aeac52f65e1a44da90aaeeab84360441'
CF.Key.set(Key)

connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = [0 for i in range(150)]

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces')
for filename in os.listdir(directory):
    if filename.endswith(".jpg"):
        #imgurl = urllib.pathname2url(os.path.join(directory, filename))
        imgurl = "C:\Users\Mit Patel\Desktop\New folder\Autoattendance-Cognitive\Cropped_faces\\face1.jpg"
        res = CF.face.detect(imgurl)
        if len(res) != 1:
            print "No face detected."
            continue

        faceIds = []
        for face in res:
            faceIds.append(face['faceId'])
        res = CF.face.identify(faceIds, personGroupId)
        print filename
        for face in res:
            if not face['candidates']:
                print "Unknown"
            else:
                personId = face['candidates'][0]['personId']
                c.execute("SELECT * FROM Students WHERE personID = ?", (personId,))
                print personId
                row = c.fetchone()
                print row
                attend[int(row[0])] += 1
                print row[1] + " recognized"
        time.sleep(6)

sheet = wb.active

columns = list(sheet)

for row in range(2,len(columns)):
 #   rn = sheet.cell('A%s' % row).value
    print row
    rn = columns[row][0].value
    if rn is not None:
        print rn;
        rn = rn[-2:]
        print rn;
        if attend[int(rn)] != 0:
            col = 2
            sheet.cell(row=row+1, column=col+1).value = 1

wb.save(filename="reports.xlsx")

#currentDir = os.path.dirname(os.path.abspath(__file__))
# imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
# res = CF.face.detect(imgurl)
# faceIds = []
# for face in res:
#   faceIds.append(face['faceId'])

# res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
# print res
